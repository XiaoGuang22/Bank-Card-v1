import os
import torch
from torch.utils.data import Dataset
from PIL import Image
import numpy as np

class CardDataset(Dataset):
    # 字符与标签映射
    CHARS = '0123456789/'
    # CHARS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    CHAR2LABEL = {char: i + 1 for i, char in enumerate(CHARS)}
    LABEL2CHAR = {label: char for char, label in CHAR2LABEL.items()}

    def __init__(self, image_dir, mode, img_height, img_width):
        texts = []
        self.mode = mode
        self.image_dir = image_dir
        self.img_height = img_height
        if mode in ("train", "val", "test"):
            file_names, texts = self._load_from_excel()
            self.file_names = file_names
        self.texts = texts

    def _load_from_excel(self):
        # 统一从 {mode}_labels.xlsx 读取
        candidate_names = [f"{self.mode}_labels.xlsx"]

        xlsx_path = None
        for name in candidate_names:
            path_try = os.path.join(self.image_dir, name)
            if os.path.isfile(path_try):
                xlsx_path = path_try
                break
        if xlsx_path is None:
            raise FileNotFoundError(f"未找到标签文件于 {self.image_dir}: 尝试 {candidate_names}")
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError("需要 openpyxl 以从 xlsx 读取标签，请先安装: pip install openpyxl") from e

        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        header_map = {cell.value: idx for idx, cell in enumerate(header_row)}
        if 'filename' not in header_map or 'CardNumberlabel' not in header_map:
            raise ValueError("xlsx 缺少必要列: 'filename' 与 'CardNumberlabel'")

        file_names = []
        texts = []
        for row in ws.iter_rows(min_row=2):
            fn_cell = row[header_map['filename']].value
            lbl_cell = row[header_map['CardNumberlabel']].value
            if fn_cell is None or lbl_cell is None:
                continue
            file_name = str(fn_cell).strip()
            label = str(lbl_cell).strip()
            # 可选：仅收录存在于目录中的文件
            img_path = os.path.join(self.image_dir, file_name)
            if os.path.isfile(img_path):
                file_names.append(file_name)
                texts.append(label)

        if len(file_names) == 0:
            raise ValueError(f"在 {xlsx_path} 中未解析到有效的 (filename, CardNumberlabel) 记录")
        return file_names, texts

    def __len__(self):
        if self.mode == "pred":
            return 1
        else:
            return len(self.file_names)

    def __getitem__(self, index):
        if self.mode in ("train", "val", "test"):
            file_name = self.file_names[index]
            file_path = os.path.join(self.image_dir, file_name)
            image = Image.open(file_path)
        elif self.mode == "pred":
            image = self.image_dir

        # 只定高，宽等比例缩放
        orig_w, orig_h = image.size
        new_w = int(orig_w * self.img_height / orig_h)
        # 如果你的模型有宽度要求（如4的倍数），可以加下面一行
        new_w = (new_w + 3) // 4 * 4
        image = image.convert('L').resize((new_w, self.img_height), Image.BILINEAR)
        image = np.array(image).reshape((1, self.img_height, new_w))
        image = (image / 127.5) - 1.0
        image = torch.FloatTensor(image)

        if len(self.texts) != 0:
            text = self.texts[index]
            target = [self.CHAR2LABEL[c] for c in text]
            target_length = [len(target)]
            target = torch.LongTensor(target)
            target_length = torch.LongTensor(target_length)
            return image, target, target_length
        else:
            return image

# 动态宽度padding的collate_fn
def dynamic_pad_collate_fn(batch):
    images, targets, target_lengths = zip(*batch)
    max_w = max(img.shape[2] for img in images)
    padded_images = []
    for img in images:
        c, h, w = img.shape
        pad_w = max_w - w
        padded_img = torch.nn.functional.pad(img, (0, pad_w), value=-1.0)
        padded_images.append(padded_img)
    images = torch.stack(padded_images)
    targets = torch.cat(targets, 0)
    target_lengths = torch.cat(target_lengths, 0)
    return images, targets, target_lengths